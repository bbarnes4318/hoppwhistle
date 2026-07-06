import os
import sys
import json
import datetime
import openpyxl
from openpyxl.utils import get_column_letter

# Dynamically resolve path to the Excel file
excel_path = "/app/rating (9) (1).xlsx"
if not os.path.exists(excel_path):
    excel_path = r"C:\Users\jimbo\OneDrive\Documents\hopbot\rating (9) (1).xlsx"

if not os.path.exists(excel_path):
    print(json.dumps({"error": f"Excel file not found at {excel_path}"}))
    sys.exit(1)

# Helper to normalize phone number to 10 digits
def clean_phone(phone_str):
    if not phone_str:
        return ""
    cleaned = "".join(c for c in str(phone_str) if c.isdigit())
    if len(cleaned) > 10 and cleaned.startswith("1"):
        cleaned = cleaned[1:]
    return cleaned[:10]

# Parse dates helper
def parse_date(date_str):
    if not date_str:
        return None
    for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d'):
        try:
            return datetime.datetime.strptime(str(date_str).split()[0], fmt).date()
        except ValueError:
            continue
    return None

def main():
    try:
        # Load leads from stdin
        input_data = sys.stdin.read()
        if not input_data.strip():
            print(json.dumps({"error": "No input data received"}))
            return
            
        leads = json.loads(input_data)
        if not isinstance(leads, list):
            leads = [leads]
            
        # Open workbook (writeable)
        wb = openpyxl.load_workbook(excel_path, data_only=False)
        sheet_calc = wb['calculations']
        sheet_key = wb['Key']
        
        # Load carrier sheets
        carriers = [
            'aflac', 'aflac-modified', 'SBLI-Modified', 'CICA', 'SBLI', 'CICA-GI', 'gtl',
            'TransAmerica', 'TransAmerica Graded', 'Corebridge', 'amam', 'amam-graded',
            'amam-return-or-premium', 'AHL', 'AHL-Graded', 'Royal Neighbors', 'Royal Neighbors-Graded',
            'Gerber-GI', 'Mutual of Omaha', 'Mutual of Omaha-Graded'
        ]
        
        carrier_data = {}
        for c in carriers:
            sheet_name = c
            # Load rates in memory as list of dicts/tuples for fast lookup
            c_sheet = wb[sheet_name]
            headers = [cell.value for cell in c_sheet[1]]
            rows = []
            for r in range(2, c_sheet.max_row + 1):
                row_vals = [cell.value for cell in c_sheet[r]]
                if any(v is not None for v in row_vals):
                    rows.append(row_vals)
            carrier_data[c] = {"headers": headers, "rows": rows}
            
        # Read Key sheet for fee and monthly_factor
        key_headers = [cell.value for cell in sheet_key[1]]
        key_rows = {}
        for r in range(2, sheet_key.max_row + 1):
            row_vals = [cell.value for cell in sheet_key[r]]
            if row_vals and row_vals[0]:
                key_rows[row_vals[0]] = row_vals
                
        def get_key_value(carrier, row_name):
            col_name = carrier
            if carrier == 'Mutual of Omaha':
                col_name = 'Mutal of Omaha'
            try:
                col_idx = key_headers.index(col_name)
                val = key_rows[row_name][col_idx]
                return float(val) if val is not None else 0.0
            except Exception:
                return 0.0
                
        # Build map of phone numbers in calculations sheet
        phone_col_idx = 3 # Column C (1-based is 3)
        existing_phones = {} # phone -> row_number
        for r in range(2, sheet_calc.max_row + 1):
            phone_val = sheet_calc.cell(row=r, column=phone_col_idx).value
            cphone = clean_phone(phone_val)
            if cphone:
                existing_phones[cphone] = r

        results = []
        excel_modified = False
        
        for lead in leads:
            phone = clean_phone(lead.get('phone', ''))
            if not phone:
                continue
                
            # Try to get Dob/BirthDate
            dob_str = lead.get('birthDate') or lead.get('dob')
            dob_date = parse_date(dob_str)
            
            # Premium Date
            prem_str = lead.get('firstPremiumDate') or lead.get('1st premium date') or "2025-11-20"
            prem_date = parse_date(prem_str) or datetime.date(2025, 11, 20)
            
            # Recalculate Age Nearest
            age = lead.get('age')
            if dob_date and prem_date:
                age = int(round((prem_date - dob_date).days / 365.25))
            else:
                try:
                    age = int(age) if age is not None else 65
                except ValueError:
                    age = 65
                    
            gender = str(lead.get('gender') or 'Male')
            tobacco = str(lead.get('smoker') or lead.get('tobacco') or 'No')
            coverage = lead.get('coverageAmount') or lead.get('coverage') or lead.get('faceAmount') or 5000
            try:
                coverage = float(coverage)
            except ValueError:
                coverage = 5000.0
                
            current_premium = lead.get('monthlyPremium') or lead.get('currentPremium') or 50.0
            try:
                current_premium = float(current_premium)
            except ValueError:
                current_premium = 50.0
                
            # Perform carrier calculations in Python
            quotes = {}
            for c in carriers:
                table = carrier_data[c]
                headers = table["headers"]
                rows = table["rows"]
                
                # Find matching row for age
                rate_row = None
                for r_vals in rows:
                    if r_vals and int(r_vals[0]) == age:
                        rate_row = r_vals
                        break
                        
                if rate_row is None:
                    continue
                    
                gender_lower = gender.lower().strip()
                tobacco_lower = tobacco.lower().strip()
                
                rate = None
                if len(headers) == 5:
                    is_smoker = tobacco_lower == 'yes'
                    if gender_lower == 'male':
                        col_lbl = 'Male Smoker' if is_smoker else 'Male Non-Smoker'
                    else:
                        col_lbl = 'Female Smoker' if is_smoker else 'Female Non-Smoker'
                    
                    for h_idx, h in enumerate(headers):
                        if h and h.replace('\xa0', ' ').strip().lower() == col_lbl.lower():
                            rate = rate_row[h_idx]
                            break
                else:
                    col_lbl = 'Male' if gender_lower == 'male' else 'Female'
                    for h_idx, h in enumerate(headers):
                        if h and h.strip().lower() == col_lbl.lower():
                            rate = rate_row[h_idx]
                            break
                            
                if rate is not None:
                    try:
                        rate = float(rate)
                        base = rate * (coverage / 1000)
                        fee = get_key_value(c, 'fee')
                        mf = get_key_value(c, 'monthly_factor')
                        if mf > 0:
                            quotes[c] = round((base + fee) * mf, 2)
                        else:
                            quotes[c] = round((base + fee) / 12, 2)
                    except ValueError:
                        pass

            # Calculate stats
            amam_quote = quotes.get('amam')
            amam_less = round(current_premium - amam_quote, 2) if amam_quote and amam_quote < current_premium else None
            
            gtl_quote = quotes.get('gtl')
            gtl_less = round(current_premium - gtl_quote, 2) if gtl_quote and gtl_quote < current_premium else None
            
            valid_quotes = {k: v for k, v in quotes.items() if v < current_premium}
            if valid_quotes:
                cheapest_carrier = min(valid_quotes, key=valid_quotes.get)
                savings = round(current_premium - valid_quotes[cheapest_carrier], 2)
            else:
                cheapest_carrier = ""
                savings = ""
                
            lead_result = {
                "firstName": lead.get('firstName', ''),
                "lastName": lead.get('lastName', ''),
                "phone": phone,
                "age": age,
                "gender": gender,
                "coverageAmount": coverage,
                "monthlyPremium": current_premium,
                "carrier": lead.get('carrier', ''),
                "primaryBeneficiaryName": lead.get('primaryBeneficiaryName', ''),
                "primaryBeneficiaryRelationship": lead.get('primaryBeneficiaryRelationship') or lead.get('relation') or '',
                "amamQuote": amam_quote,
                "amamLessThanCurrent": amam_less,
                "gtlQuote": gtl_quote,
                "gtlLessThanCurrent": gtl_less,
                "cheapestCarrierUnderCurrent": cheapest_carrier,
                "savingsVsCurrent": savings,
                "dob": dob_date.strftime('%Y-%m-%d') if dob_date else None,
                "firstPremiumDate": prem_date.strftime('%Y-%m-%d') if prem_date else None,
            }
            results.append(lead_result)
            
            # Update Excel workbook
            target_row = existing_phones.get(phone)
            if not target_row:
                target_row = sheet_calc.max_row + 1
                existing_phones[phone] = target_row
                
            # Map lead fields into cells
            sheet_calc.cell(row=target_row, column=1, value=lead.get('firstName', ''))
            sheet_calc.cell(row=target_row, column=2, value=lead.get('lastName', ''))
            sheet_calc.cell(row=target_row, column=3, value=phone)
            sheet_calc.cell(row=target_row, column=4, value=age)
            if dob_date:
                sheet_calc.cell(row=target_row, column=5, value=dob_date.strftime('%Y-%m-%d'))
            sheet_calc.cell(row=target_row, column=6, value=gender)
            sheet_calc.cell(row=target_row, column=7, value=lead.get('address', ''))
            sheet_calc.cell(row=target_row, column=8, value=lead.get('state', ''))
            sheet_calc.cell(row=target_row, column=9, value=lead.get('height', ''))
            sheet_calc.cell(row=target_row, column=10, value=lead.get('weight', ''))
            sheet_calc.cell(row=target_row, column=11, value=tobacco)
            sheet_calc.cell(row=target_row, column=12, value=lead.get('email', ''))
            sheet_calc.cell(row=target_row, column=13, value=lead.get('burialCremation', ''))
            sheet_calc.cell(row=target_row, column=14, value=lead.get('primaryBeneficiaryName', ''))
            sheet_calc.cell(row=target_row, column=15, value=lead_result['primaryBeneficiaryRelationship'])
            sheet_calc.cell(row=target_row, column=18, value=lead.get('carrier', ''))
            sheet_calc.cell(row=target_row, column=19, value=coverage)
            sheet_calc.cell(row=target_row, column=20, value=current_premium)
            if prem_date:
                sheet_calc.cell(row=target_row, column=21, value=prem_date.strftime('%Y-%m-%d'))
            sheet_calc.cell(row=target_row, column=22, value=lead.get('monthlyRecurringDueDate', ''))
            sheet_calc.cell(row=target_row, column=23, value=lead.get('ssn', ''))
            sheet_calc.cell(row=target_row, column=24, value=lead.get('driversLicense', ''))
            sheet_calc.cell(row=target_row, column=25, value=lead.get('bankName', ''))
            sheet_calc.cell(row=target_row, column=26, value=lead.get('routingNumber', ''))
            sheet_calc.cell(row=target_row, column=27, value=lead.get('accountNumber', ''))
            sheet_calc.cell(row=target_row, column=28, value=lead.get('health', ''))
            sheet_calc.cell(row=target_row, column=29, value=lead.get('medications', ''))
            sheet_calc.cell(row=target_row, column=30, value=lead.get('doctorName', ''))
            
            # Map age again in Col AE
            sheet_calc.cell(row=target_row, column=31, value=age)
            
            # Map quotes in columns AF to AY
            for col_idx in range(32, 52):
                col_let = get_column_letter(col_idx)
                header_val = sheet_calc.cell(row=1, column=col_idx).value
                # Match header to quotes keys
                for c in carriers:
                    if c.lower() in header_val.lower():
                        sheet_calc.cell(row=target_row, column=col_idx, value=quotes.get(c))
                        break
                        
            # Map AZ to BE quotes
            sheet_calc.cell(row=target_row, column=52, value=amam_quote) # AZ
            sheet_calc.cell(row=target_row, column=53, value=amam_less)  # BA
            sheet_calc.cell(row=target_row, column=54, value=gtl_quote)  # BB
            sheet_calc.cell(row=target_row, column=55, value=gtl_less)   # BC
            sheet_calc.cell(row=target_row, column=56, value=cheapest_carrier) # BD
            sheet_calc.cell(row=target_row, column=57, value=savings)    # BE
            
            excel_modified = True
            
        if excel_modified:
            wb.save(excel_path)
            
        print(json.dumps(results))
        
    except Exception as e:
        print(json.dumps({"error": str(e)}))
        sys.exit(1)

if __name__ == "__main__":
    main()
